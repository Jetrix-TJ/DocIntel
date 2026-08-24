"""Detect hidden rows/columns/sheets in an XLSX that carry non-empty data.

Not an extraction path - this never reads a value as a fact, only asks "is
there something here a human should look at that no render could ever show?"
A hidden cell is, by definition, excluded from any visual layout: no render -
LibreOffice's or anyone else's - can ever display it (a measured finding from
this project's XLSX-vs-render research spike: a hidden reconciliation column
survived a native `openpyxl` read but is structurally invisible to the
render-to-PDF path this project's XLSX documents otherwise rely on). Rather
than guess at what a hidden value means, or build a whole second extraction
engine to read it "correctly", this follows the same posture
`extract.annotations.detect_flattened` already established for a different
structurally-undetectable-by-render signal (flattened human markup): detect
it, tag it, and let a human decide - never silently trust it, never silently
ignore it.

`openpyxl` is lazily imported, matching this project's existing optional-
dependency convention (see `export/excel.py`). It is not required to process
an XLSX document at all today (LibreOffice does the actual PDF conversion);
this check is the first thing that needs it for XLSX INPUT, so its absence is
treated the same way `.msg` support treats a missing `extract-msg` package
(`adapters/intake/email.py::_require_extract_msg`) - a loud, actionable,
per-document `PermanentError`, not a silently-skipped check. A safety net
whose own absence is silent would defeat the reason it exists.
"""

from __future__ import annotations

from docintel.core.errors import PermanentError


def _require_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise PermanentError(
            "checking an XLSX document for hidden content needs the optional "
            "'openpyxl' package - pip install 'docintel[export]'"
        ) from exc
    return openpyxl


def has_hidden_content(path: str) -> bool:
    """True if any hidden sheet, row, or column in the workbook at `path`
    contains at least one non-empty cell.

    A hidden sheet's own cells all count, regardless of that sheet's row/
    column visibility - if the sheet itself is hidden, none of it can ever
    render. Checked cheaply: this stops at the first non-empty hidden cell
    found, since the answer only needs to be "yes, look at this" or "no
    hidden content at all" - not an inventory of what's hidden.

    Raises `PermanentError` for anything `openpyxl` cannot open - the same
    discipline `extract.convert`'s converters already follow, never letting a
    raw library exception surface uncaught.
    """
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise PermanentError(
            f"could not open {path!r} as an XLSX workbook to check for hidden content: {exc}"
        ) from exc

    try:
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                if _any_non_empty_cell(ws):
                    return True
                continue
            if _hidden_row_or_column_has_data(ws):
                return True
        return False
    finally:
        wb.close()


def _any_non_empty_cell(ws) -> bool:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                return True
    return False


def _hidden_row_or_column_has_data(ws) -> bool:
    hidden_columns = {letter for letter, dim in ws.column_dimensions.items() if dim.hidden}
    hidden_rows = {index for index, dim in ws.row_dimensions.items() if dim.hidden}
    if not hidden_columns and not hidden_rows:
        return False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue
            if cell.row in hidden_rows or cell.column_letter in hidden_columns:
                return True
    return False
