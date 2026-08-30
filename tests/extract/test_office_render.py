"""`extract.office_render`: tier-2 LibreOffice-free XLSX fallback - renders a
real image from the original workbook, for `pipeline.stages.s5b_vision` to
hand to a vision model exactly like any other image.
"""

from __future__ import annotations

import openpyxl
import pytest
from PIL import Image

from docintel.extract import office_render


def _workbook(path, rows, hidden_columns=(), hidden_rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    for letter in hidden_columns:
        ws.column_dimensions[letter].hidden = True
    for index in hidden_rows:
        ws.row_dimensions[index].hidden = True
    wb.save(path)
    return path


def test_xlsx_to_image_produces_a_valid_nonempty_png(tmp_path):
    xlsx = tmp_path / "invoice.xlsx"
    _workbook(xlsx, [["Vendor Name: ACME CORP"], ["Total Due: $500.00"]])

    out_path = office_render.xlsx_to_image(str(xlsx))

    assert out_path.endswith(".png")
    img = Image.open(out_path)
    img.load()
    assert img.width > 0
    assert img.height > 0


def test_visible_rows_excludes_hidden_column(tmp_path):
    xlsx = tmp_path / "hidden-col.xlsx"
    _workbook(
        xlsx,
        [["Vendor Name: ACME CORP", "internal reconciled total 999.99"]],
        hidden_columns=["B"],
    )
    openpyxl_wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    rows = office_render._visible_rows(openpyxl_wb)

    flat = " ".join(v for row in rows for v in row)
    assert "ACME CORP" in flat
    assert "999.99" not in flat


def test_visible_rows_excludes_hidden_row(tmp_path):
    xlsx = tmp_path / "hidden-row.xlsx"
    _workbook(
        xlsx,
        [["Vendor Name: ACME CORP"], ["internal note 999.99"]],
        hidden_rows=[2],
    )
    openpyxl_wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    rows = office_render._visible_rows(openpyxl_wb)

    flat = " ".join(v for row in rows for v in row)
    assert "ACME CORP" in flat
    assert "999.99" not in flat


def test_missing_openpyxl_raises_permanent_error(tmp_path, monkeypatch):
    from docintel.core.errors import PermanentError

    monkeypatch.setattr(office_render, "_require_openpyxl", lambda: (_ for _ in ()).throw(
        PermanentError("openpyxl missing")
    ))
    with pytest.raises(PermanentError):
        office_render.xlsx_to_image(str(tmp_path / "does-not-matter.xlsx"))


def test_xlsx_to_image_raises_permanent_error_over_the_row_cap(tmp_path, monkeypatch):
    from docintel.core.errors import PermanentError

    monkeypatch.setattr(office_render, "_MAX_ROWS", 2)
    xlsx = tmp_path / "too-many-rows.xlsx"
    _workbook(xlsx, [["Row 1"], ["Row 2"], ["Row 3"]])

    with pytest.raises(PermanentError, match="3 visible populated rows"):
        office_render.xlsx_to_image(str(xlsx))
