"""`extract.office_fallback`: tier-1 LibreOffice-free XLSX rendering - a real
`.html` table `extract.plaintext.load_document` can read completely
unmodified.
"""

from __future__ import annotations

import openpyxl
import pytest

from docintel.extract import office_fallback, plaintext


def _workbook(path, rows, hidden_columns=(), hidden_rows=()):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    for letter in hidden_columns:
        ws.column_dimensions[letter].hidden = True
    for index in hidden_rows:
        ws.row_dimensions[index].hidden = True
    wb.save(path)
    return path


def test_visible_cells_roundtrip_through_the_native_html_reader(tmp_path):
    xlsx = tmp_path / "invoice.xlsx"
    _workbook(xlsx, [
        ["Vendor Name: ACME CORP"],
        ["Invoice Number: INV-1001"],
        ["Total Due: $500.00"],
    ])

    html_path = office_fallback.xlsx_to_html(str(xlsx))

    pages, page_meta, text_source = plaintext.load_document(html_path, ".html")
    full_text = " ".join(word.text for word in pages[0].words)
    assert "ACME CORP" in full_text
    assert "INV-1001" in full_text
    assert "500.00" in full_text
    assert text_source == "native"


def test_hidden_column_is_excluded(tmp_path):
    xlsx = tmp_path / "hidden-col.xlsx"
    _workbook(
        xlsx,
        [["Vendor Name: ACME CORP", "internal reconciled total 999.99"]],
        hidden_columns=["B"],
    )

    html_path = office_fallback.xlsx_to_html(str(xlsx))

    pages, _, _ = plaintext.load_document(html_path, ".html")
    full_text = " ".join(word.text for word in pages[0].words)
    assert "ACME CORP" in full_text
    assert "999.99" not in full_text


def test_hidden_row_is_excluded(tmp_path):
    xlsx = tmp_path / "hidden-row.xlsx"
    _workbook(
        xlsx,
        [["Vendor Name: ACME CORP"], ["internal note 999.99"]],
        hidden_rows=[2],
    )

    html_path = office_fallback.xlsx_to_html(str(xlsx))

    pages, _, _ = plaintext.load_document(html_path, ".html")
    full_text = " ".join(word.text for word in pages[0].words)
    assert "ACME CORP" in full_text
    assert "999.99" not in full_text


def test_hidden_sheet_is_excluded_entirely(tmp_path):
    xlsx = tmp_path / "hidden-sheet.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Visible"
    wb.active["A1"] = "Vendor Name: ACME CORP"
    hidden_ws = wb.create_sheet("Hidden")
    hidden_ws["A1"] = "internal note 999.99"
    hidden_ws.sheet_state = "hidden"
    wb.save(xlsx)

    html_path = office_fallback.xlsx_to_html(str(xlsx))

    pages, _, _ = plaintext.load_document(html_path, ".html")
    full_text = " ".join(word.text for word in pages[0].words)
    assert "ACME CORP" in full_text
    assert "999.99" not in full_text


def test_missing_openpyxl_raises_permanent_error(tmp_path, monkeypatch):
    from docintel.core.errors import PermanentError

    monkeypatch.setattr(office_fallback, "_require_openpyxl", lambda: (_ for _ in ()).throw(
        PermanentError("openpyxl missing")
    ))
    with pytest.raises(PermanentError):
        office_fallback.xlsx_to_html(str(tmp_path / "does-not-matter.xlsx"))
