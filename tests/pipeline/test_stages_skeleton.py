from northstar import PACK as NORTHSTAR_PACK

from docintel.adapters.vision.fake import FakeVision
from docintel.core.contract import validate_record
from docintel.core.models import new_context
from docintel.pipeline.hooks import HookRegistry
from docintel.pipeline.runner import Runner
from docintel.pipeline.stages import build_default_stages, build_pipeline

CORPUS = "docs/_AP Invoice 6060DTSS        D.T.S.S. Inc. 699.00000.pdf"


def _runner():
    return Runner(stages=build_default_stages(vision=FakeVision()), hooks=HookRegistry())


def test_every_stage_runs_and_is_logged():
    rec = _runner().process("d1", CORPUS)
    validate_record(rec)


def test_the_default_sequence_is_eleven_modules_in_pipeline_order():
    names = [s.name for s in build_default_stages(vision=FakeVision())]
    assert names == [
        "intake", "attachment_filter", "classify", "persona_lookup",
        "resolve_processing_profile", "apply_cached_rules", "vision_one_shot",
        "agent_escalation", "capture_fields", "confidence_gate", "emit_record",
    ]


def test_every_stage_records_that_it_ran():
    """'Pass any PDF and it traverses all 8 stages' must be OBSERVABLY true.

    Asserts on the event log itself, which is the only evidence a stage ran.
    5a is absent by design here: no persona exists, so the document is a hard
    miss and routes to 5b.
    """
    captured = {}

    class Spy:
        name = "spy"

        def run(self, ctx):
            captured["events"] = list(ctx.events)
            return ctx

    stages = build_default_stages(vision=FakeVision())
    stages.append(Spy())
    Runner(stages=stages, hooks=HookRegistry()).process("d1", CORPUS)

    log = " ".join(captured["events"])
    for marker in ("s1:", "s2:", "s3:", "s4:", "s5b:", "s6:", "s7:", "s8:"):
        assert marker in log, f"no evidence stage {marker} ran; log was {log!r}"


def test_hard_miss_routes_to_vision_not_cached_rules():
    """Skeleton has no personas, so every document is a hard miss -> 5b."""
    rec = _runner().process("d1", CORPUS)
    assert rec["extraction_route"] == "5b_vision"


class _StubPersona:
    rule_version = "v14"


class _StubStore:
    """Stands in for the Persona DB, which arrives in cluster C7."""

    def __init__(self, persona: object | None) -> None:
        self.persona = persona

    def lookup(self, fingerprint: str, doc_type: str | None) -> object | None:
        return self.persona


class _StubExecutor:
    """Stands in for the grammar executor, which arrives in cluster C2."""

    def __init__(self, quality: float) -> None:
        self.quality = quality

    def apply(self, ctx):
        ctx.extracted.set("invoice_number", "AC-002561", self.quality)
        ctx.extracted.set("total_printed", "1284.50", self.quality)
        return ctx


def _routing_runner(persona, quality, vision):
    """A stage list wired for one specific stage-5 routing path."""
    from docintel.pipeline.stages.s1_intake import Intake
    from docintel.pipeline.stages.s2_filter import AttachmentFilter
    from docintel.pipeline.stages.s3_classify import Classify
    from docintel.pipeline.stages.s4_persona import PersonaLookup
    from docintel.pipeline.stages.s5a_cached import ApplyCachedRules
    from docintel.pipeline.stages.s5b_vision import VisionOneShot
    from docintel.pipeline.stages.s5c_agent import AgentEscalation
    from docintel.pipeline.stages.s6_capture import CaptureFields
    from docintel.pipeline.stages.s7_gate import ConfidenceGate
    from docintel.pipeline.stages.s8_emit import EmitRecord

    return Runner(
        stages=[
            Intake(), AttachmentFilter(), Classify(),
            PersonaLookup(store=_StubStore(persona)),
            ApplyCachedRules(executor_factory=lambda persona: _StubExecutor(quality)),
            VisionOneShot(vision=vision), AgentEscalation(),
            CaptureFields(), ConfidenceGate(), EmitRecord(),
        ],
        hooks=HookRegistry(),
    )


def test_persona_hit_with_good_confidence_takes_the_fast_lane_with_zero_vision_calls():
    """The economics of the whole design: a persona hit must cost no AI call."""
    vision = FakeVision()
    rec = _routing_runner(_StubPersona(), quality=0.95, vision=vision).process("d1", CORPUS)
    assert rec["extraction_route"] == "5a_cached"
    assert vision.calls == [], "the fast lane must make ZERO vision calls"
    assert rec["extraction_rule_version"] == "v14"


def test_persona_hit_whose_rules_collapse_falls_back_to_vision():
    """Old selectors against a redesigned template: emit trustworthy values anyway."""
    vision = FakeVision()
    rec = _routing_runner(_StubPersona(), quality=0.10, vision=vision).process("d1", CORPUS)
    assert vision.calls != [], "a collapsed persona must fall back to the vision one-shot"
    assert rec["extraction_route"] == "5b_vision"


def test_soft_miss_still_runs_the_cached_rules_first():
    """Layout drift is usually cosmetic, so try the rules before paying for vision."""
    from docintel.pipeline.stages.s5a_cached import ApplyCachedRules

    ctx = new_context(document_id="d1", source_path=CORPUS)
    ctx.persona_status = "soft_miss"
    ctx.persona = _StubPersona()
    out = ApplyCachedRules(
        executor_factory=lambda persona: _StubExecutor(0.95)
    ).run(ctx)
    assert out.extraction_route == "5a_cached"
    assert out.extracted.get("invoice_number") == "AC-002561"


def test_hard_miss_sets_review_not_regen():
    """A first-time sender has no rules, so regen_flag would be meaningless."""
    rec = _runner().process("d1", CORPUS)
    assert rec["review_flag"] is True
    assert rec["regen_flag"] is False, (
        "regen_flag means 'the rules are wrong'; a hard miss has no rules. "
        "Stage 7 is the sole writer of regen_flag."
    )


def test_unsupported_file_type_is_skipped_with_a_reason_never_dropped():
    # `.pptx` specifically: still not in `ACCEPTED_SUFFIXES` today (Phase 4's
    # TXT/CSV/HTML onboarding did not add it) - `.txt` USED to be the example
    # here before it became a genuinely supported format.
    rec = _runner().process("d2", "/tmp/notes.pptx")
    validate_record(rec)
    assert rec["disposition"] == "skipped"
    assert rec["reason"]


def test_a_scanned_image_is_no_longer_skipped_and_reads_via_ocr(tmp_path):
    """The Stage 2 gap this closes: before `extract.convert` existed, this
    exact PNG would have hit the same `not in ALLOWED_SUFFIXES` branch as
    `.txt` above. Now it converts, and the wrapped page - no text layer at
    all - takes the OCR path exactly like a scanned PDF always has."""
    from PIL import Image

    png = tmp_path / "scan.png"
    Image.new("RGB", (850, 1100), (255, 255, 255)).save(png)

    rec = _runner().process("d-img", str(png))

    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert rec["text_source"] == "ocr"


def test_a_txt_document_is_never_converted_ocrd_or_given_a_temp_directory(tmp_path, monkeypatch):
    """Phase 4: TXT/CSV/HTML never touch `extract.convert` or `extract.ocr`
    at all - not even lazily at vision, since there is no persona for this
    document and it will reach Stage 5b, which must still never call either
    module for a plaintext source."""
    import tempfile

    txt = tmp_path / "invoice.txt"
    txt.write_text(
        "Invoice Number: INV-77001\nTotal Due: 512.00\nAccount: 00812-QR\n",
        encoding="utf-8",
    )

    created: list[str] = []
    real_mkdtemp = tempfile.mkdtemp

    def spy_mkdtemp(*args, **kwargs):
        directory = real_mkdtemp(*args, **kwargs)
        created.append(directory)
        return directory

    monkeypatch.setattr(tempfile, "mkdtemp", spy_mkdtemp)

    rec = _runner().process("d-txt-native", str(txt))

    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert rec["text_source"] == "native"
    assert created == [], "a TXT document must never trigger any conversion"


def test_a_csv_document_reaches_the_end_of_the_pipeline(tmp_path):
    csv_path = tmp_path / "invoice.csv"
    csv_path.write_text(
        "Vendor,Invoice Number,Total\nACME Utilities,INV-77002,640.00\n",
        encoding="utf-8",
    )
    rec = _runner().process("d-csv-native", str(csv_path))
    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert rec["text_source"] == "native"


def test_an_html_document_reaches_the_end_of_the_pipeline(tmp_path):
    html_path = tmp_path / "invoice.html"
    html_path.write_text(
        "<html><body><p>Invoice Number: INV-77003</p><p>Total Due: 780.00</p></body></html>",
        encoding="utf-8",
    )
    rec = _runner().process("d-html-native", str(html_path))
    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert rec["text_source"] == "native"


def test_a_gemini_native_image_never_creates_a_temp_directory(tmp_path, monkeypatch):
    """PNG/JPG are never converted to PDF at any stage - not eagerly at
    Stage 2 (OCR and annotation detection read the source bytes directly),
    and not lazily at Stage 5b either (Gemini understands them natively,
    verified against live `ai.google.dev` docs). No conversion means no
    `mkdtemp()` call at all, for a document that has no persona and
    therefore does reach vision."""
    import tempfile

    from PIL import Image

    png = tmp_path / "scan.png"
    Image.new("RGB", (850, 1100), (255, 255, 255)).save(png)

    created: list[str] = []
    real_mkdtemp = tempfile.mkdtemp

    def spy_mkdtemp(*args, **kwargs):
        directory = real_mkdtemp(*args, **kwargs)
        created.append(directory)
        return directory

    monkeypatch.setattr(tempfile, "mkdtemp", spy_mkdtemp)

    rec = _runner().process("d-png-no-conversion", str(png))

    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert created == [], "a Gemini-native image must never be converted to PDF"


def test_a_non_gemini_native_image_converts_lazily_only_at_vision_and_cleans_up(
    tmp_path, monkeypatch
):
    """TIFF is Pillow-native for OCR/annotation detection (no conversion
    needed at Stage 2 for those), but is NOT a Gemini-documented image MIME
    type, so it still needs to become a PDF before vision can read it. That
    conversion must happen lazily, in Stage 5b, only for a document that
    actually reaches vision (no persona, here) - never eagerly at Stage 2 -
    and the Runner must still remove the resulting temp directory once this
    document's whole run is over."""
    import os
    import tempfile

    from PIL import Image

    tiff = tmp_path / "scan.tiff"
    Image.new("RGB", (850, 1100), (255, 255, 255)).save(tiff)

    created: list[str] = []
    real_mkdtemp = tempfile.mkdtemp

    def spy_mkdtemp(*args, **kwargs):
        directory = real_mkdtemp(*args, **kwargs)
        created.append(directory)
        return directory

    monkeypatch.setattr(tempfile, "mkdtemp", spy_mkdtemp)

    rec = _runner().process("d-tiff-lazy-cleanup", str(tiff))

    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert created, "a TIFF reaching vision must still convert to PDF, lazily"
    assert all(not os.path.isdir(d) for d in created), (
        "every temp dir the lazy conversion created must be removed once processing is done"
    )


def test_a_soft_fingerprint_duplicate_is_flagged_when_no_invoice_number_is_extracted(tmp_path):
    """The gap this closes: a hard-miss/collapsed document with no invoice
    number and no account+period previously could never be flagged as a
    possible duplicate, even byte-for-byte identical - `document_identity`
    fell all the way through to `None`, and `IdentityIndex` no-ops on `None`.
    Reprocessing the exact same file twice (same sender, filename, byte size)
    now flags the second as a possible duplicate of the first via Stage 1's
    `soft_fingerprint` - the weakest identity rung, but not no rung at all."""
    from PIL import Image

    png = tmp_path / "scan.png"
    Image.new("RGB", (850, 1100), (255, 255, 255)).save(png)

    runner = _runner()
    first = runner.process("dup-a", str(png), sender_email="ap@acme.example")
    second = runner.process("dup-b", str(png), sender_email="ap@acme.example")

    assert first["derived"]["identity_basis"] == "soft_fingerprint"
    assert first["possible_duplicate_of"] is None
    assert second["derived"]["identity_basis"] == "soft_fingerprint"
    assert second["possible_duplicate_of"] == "dup-a"


def test_an_xlsx_with_hidden_content_is_tagged_and_forced_to_review(tmp_path, monkeypatch):
    """Phase 3c-i: hidden content in the ORIGINAL workbook is structurally
    invisible to the rendered PDF this document also goes through - Stage 2
    must still tag it, and Stage 7 must still force review, entirely
    independent of how well the rendered-PDF path scored."""
    import openpyxl

    from docintel.extract import convert

    def fake_convert(path):
        out = tmp_path / "converted.pdf"
        from PIL import Image

        Image.new("RGB", (100, 100)).save(out, "PDF")
        return str(out)

    monkeypatch.setattr(convert, "convert_office_to_pdf", fake_convert)

    xlsx = tmp_path / "invoice.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Invoice"
    ws["C1"] = "adjusted total, never printed"
    ws.column_dimensions["C"].hidden = True
    wb.save(xlsx)

    rec = _runner().process("d-xlsx-hidden", str(xlsx))

    validate_record(rec)
    assert "xlsx_hidden_content_present" in rec["tags"]
    assert rec["lane"] == "review"
    assert rec["review_flag"] is True


def test_an_xlsx_with_no_hidden_content_is_not_tagged(tmp_path, monkeypatch):
    import openpyxl

    from docintel.extract import convert

    def fake_convert(path):
        out = tmp_path / "converted.pdf"
        from PIL import Image

        Image.new("RGB", (100, 100)).save(out, "PDF")
        return str(out)

    monkeypatch.setattr(convert, "convert_office_to_pdf", fake_convert)

    xlsx = tmp_path / "clean.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Invoice"
    wb.save(xlsx)

    rec = _runner().process("d-xlsx-clean", str(xlsx))

    validate_record(rec)
    assert "xlsx_hidden_content_present" not in rec["tags"]


def test_reprocessing_the_same_office_document_reuses_the_conversion_cache(
    tmp_path, monkeypatch
):
    """Phase 2's whole point: a retried/reprocessed DOCX must not pay
    LibreOffice's cost twice, and the Runner's unconditional per-document
    `ctx.temp_dirs` cleanup (`pipeline/runner.py`) must never reach into the
    long-lived cache directory a cache HIT is served from."""
    from docintel.extract import convert, convert_cache

    monkeypatch.setattr(convert_cache, "CACHE_DIR", tmp_path / "convert-cache")

    calls = []

    def fake_convert(path):
        # Mimics the real `convert_office_to_pdf`'s own `tempfile.mkdtemp()`
        # isolation: its output must live in a FRESH directory disjoint from
        # `convert_cache.CACHE_DIR`, exactly as the real function's output
        # (under the OS temp root) is always disjoint from the project's
        # `var/convert-cache`. A fake that put its output inside the same
        # tree `CACHE_DIR` lives under would make `os.path.dirname(real_path)`
        # an ANCESTOR of the cache directory - registering that in
        # `ctx.temp_dirs` would delete the cache itself, a bug in this test
        # fixture, not in the production code it's standing in for.
        import os
        import tempfile

        calls.append(path)
        out_dir = tempfile.mkdtemp(dir=tmp_path, prefix="fake-office-convert-")
        out = os.path.join(out_dir, "converted.pdf")
        from PIL import Image

        Image.new("RGB", (100, 100)).save(out, "PDF")
        return out

    monkeypatch.setattr(convert, "convert_office_to_pdf", fake_convert)

    docx = tmp_path / "invoice.docx"
    docx.write_bytes(b"not a real docx - the converter is faked for this test")

    runner = _runner()
    first = runner.process("d-docx-1", str(docx))
    second = runner.process("d-docx-2", str(docx))

    validate_record(first)
    validate_record(second)
    assert calls == [str(docx)], "the second run must be served from the cache, not re-converted"
    assert (tmp_path / "convert-cache").is_dir(), (
        "the cache directory must survive both documents' per-run temp_dirs cleanup"
    )
    assert list((tmp_path / "convert-cache").glob("*.pdf")), (
        "the cached PDF itself must still be on disk after both runs"
    )


def test_an_office_document_reaches_the_office_converter_not_the_allowlist_gate(
    tmp_path, monkeypatch
):
    """Proves the WIRING - that a `.docx` now reaches
    `convert.convert_office_to_pdf` instead of being rejected at the
    allowlist - without needing a real LibreOffice install. The converter
    itself is proven separately, exhaustively, in `tests/extract/test_convert.py`."""
    from docintel.extract import convert

    calls = []

    def fake_convert(path):
        calls.append(path)
        # Hand back a real, tiny PDF so the rest of the pipeline has
        # something legitimate to read.
        import pdfplumber  # noqa: F401  (import proves the fixture path exists)

        out = tmp_path / "converted.pdf"
        from PIL import Image

        Image.new("RGB", (100, 100)).save(out, "PDF")
        return str(out)

    monkeypatch.setattr(convert, "convert_office_to_pdf", fake_convert)

    docx = tmp_path / "invoice.docx"
    docx.write_bytes(b"not a real docx - the converter is faked for this test")

    rec = _runner().process("d-docx", str(docx))

    validate_record(rec)
    assert rec["disposition"] != "skipped"
    assert calls == [str(docx)]


def test_document_id_is_stable_for_the_same_source():
    r = _runner()
    a = r.process("stable-id", CORPUS)
    b = r.process("stable-id", CORPUS)
    assert a["document_id"] == b["document_id"] == "stable-id"


def test_build_pipeline_wires_one_shared_jobs_object_into_both_escalation_stages():
    """`AgentEscalation` (a hard-miss sender) and `ConfidenceGate` (an unknown
    prior_balance_basis) enqueue two different kinds of job from two different
    points in the pipeline - both must write into the SAME queue instance, or
    a reviewer's /review page would only ever see half of what's pending.
    """
    jobs = object()
    runner = build_pipeline(vision=FakeVision(), jobs=jobs)
    escalation = next(s for s in runner.stages if s.name == "agent_escalation")
    gate = next(s for s in runner.stages if s.name == "confidence_gate")
    assert escalation.jobs is jobs
    assert gate.jobs is jobs


def test_build_pipeline_defaults_jobs_to_none_not_a_real_queue():
    """`jobs` must stay a safe no-op by default, exactly like `vision`/`store`
    elsewhere in this module - a function used from tests, the CLI, and the
    web UI is the wrong place for a surprising disk side effect (a real,
    shared `var/jobs.sqlite3` opened just because a caller omitted `jobs=`).
    """
    runner = build_pipeline(vision=FakeVision())
    escalation = next(s for s in runner.stages if s.name == "agent_escalation")
    gate = next(s for s in runner.stages if s.name == "confidence_gate")
    assert escalation.jobs is None
    assert gate.jobs is None


def test_build_pipeline_defaults_to_real_retries_not_runners_own_zero():
    """`Runner.__init__`'s own `max_retries=0` default exists for direct,
    low-level construction in tests that want a `TransientError` to fail on
    the first try. `build_pipeline` is the one function every real caller
    (`cli.py::_build_runner`, `webui/app.py::create_app`) goes through, so if
    IT silently forwarded that same `0`, a transient Gemini rate-limit or a
    timed-out Office conversion would never actually get retried in
    production despite the retry machinery existing and being unit-tested in
    isolation (`tests/pipeline/test_runner.py`)."""
    runner = build_pipeline(vision=FakeVision())
    assert runner.max_retries > 0


def test_build_pipeline_max_retries_is_still_overridable():
    runner = build_pipeline(vision=FakeVision(), max_retries=5)
    assert runner.max_retries == 5


def test_build_pipeline_preserves_a_caller_supplied_hooks_registry():
    """A real-time integrator registers a `beforeEmit` hook BEFORE calling
    `build_pipeline` to get notified the instant a document needs a human
    (`ctx.review_flag`/`ctx.lane`, both set by Stage 7 before `beforeEmit`
    fires) - that hook must survive `build_pipeline`'s own wiring rather than
    being silently discarded in favor of a fresh registry.
    """
    hooks = HookRegistry()
    seen: list[tuple[str, bool, str | None]] = []

    def notify_if_needs_review(ctx, nxt):
        ctx = nxt(ctx)
        seen.append((ctx.document_id, ctx.review_flag, ctx.lane))
        return ctx

    hooks.register("beforeEmit", notify_if_needs_review, pack="my_integration")
    runner = build_pipeline(vision=FakeVision(), hooks=hooks)

    runner.process("d1", CORPUS)

    assert len(seen) == 1
    assert seen[0][0] == "d1"


def test_build_pipeline_still_registers_pack_hooks_alongside_a_caller_supplied_registry():
    """The domain packs' own hooks (`register_all`) must still land on the SAME
    registry a caller passed in - not skipped just because the registry wasn't
    freshly created here.
    """
    hooks = HookRegistry()
    build_pipeline(vision=FakeVision(), hooks=hooks, extra_packs=[NORTHSTAR_PACK])

    assert hooks.registered("classifySignals") != []


def test_build_pipeline_without_hooks_reproduces_old_behavior():
    """Omitting `hooks` must still process a document exactly as before -
    the new parameter is purely additive."""
    runner = build_pipeline(vision=FakeVision())
    rec = runner.process("d1", CORPUS)
    validate_record(rec)


def _throwaway_pack(directory):
    """A minimal, valid `DataPack` for testing `extra_packs` - not a fixture
    from disk, since the test only needs the object to exist and be present
    in `Classify.packs`, never to actually claim a real document."""
    from docintel.packs.datapack import DataPack

    spec = {
        "name": "acme_widgets",
        "doc_types": ["standard_invoice"],
        "fields": {"standard_invoice": {"all": [], "required": [], "any_of": [], "derived_only": []}},
        "claim": {
            "rules": [{"kind": "markers", "scope": "primary", "values": ["acme widgets"]}],
            "vetoes": [],
        },
        "ladder": {
            "default": "standard_invoice",
            "rungs": [{
                "name": "r", "doc_type": "standard_invoice",
                "when": {"signal": "pattern_in_scope", "params": {"pattern": "never-matches", "scope": "primary"}},
            }],
        },
    }
    return DataPack(spec, directory=str(directory))


def test_build_pipeline_appends_extra_packs_to_the_shipped_ones(tmp_path):
    """The extension point for a wholly new company no shipped pack claims at
    all (see `registry.load_extra_personas`/`load_extra_aliases` for the
    OTHER extension point: a new vendor inside an already-shipped pack).

    No pack ships by default anymore (`PACK_MODULES`/`PACK_FILES` are both
    empty), so "appends to the shipped ones" is proven against a real
    shipped-pack STAND-IN (`extra_packs` given twice, in list order) rather
    than an actual shipped pack - the mechanism under test is that
    `extra_packs` EXTENDS whatever `load_packs()` returns, never replaces it,
    which holds just as true when that list is empty.
    """
    pack = _throwaway_pack(tmp_path)

    runner = build_pipeline(vision=FakeVision(), extra_packs=[NORTHSTAR_PACK, pack])

    classify = next(s for s in runner.stages if s.name == "classify")
    assert pack in classify.packs
    assert NORTHSTAR_PACK in classify.packs
    assert list(classify.packs) == [NORTHSTAR_PACK, pack], "extra_packs order is preserved, nothing dropped"


def test_build_pipeline_extra_packs_defaults_to_no_change():
    """Omitting `extra_packs` must still process a document exactly as
    before - the new parameter is purely additive, same discipline as
    `hooks=`."""
    runner = build_pipeline(vision=FakeVision())
    rec = runner.process("d1", CORPUS)
    validate_record(rec)
